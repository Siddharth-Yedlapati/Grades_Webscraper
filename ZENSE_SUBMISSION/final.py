import time
import os
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import ElementNotInteractableException
from difflib import SequenceMatcher
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime, timedelta


options = Options()
options.headless = True
options.add_argument("window-size=1400,880")
driver = webdriver.Chrome("./chromedriver", options=options)

def scrape():
    username = input("enter your username here : ")
    password = input("enter your password here : ")

    start_date1 = input("enter starting data of query(DD/MM/YYYY): ")

    topics = []
    names = []
    recordings = []
    dates = []

    start_date = start_date1.split("/")      # converting date into a list[day, month, year]

    start_day = start_date[0]
    start_day = int(start_day)
    end_day = start_day + 1
    start_day = str(start_day)

    start_month = int(start_date[1]) - 1      # since codetantra uses month indexing of 0 to 11, 1 is subtracted from the start month
    start_year = start_date[2]

    end_day = str(end_day)
    end_year = start_date[2]

    now = datetime.now()
    year = now.strftime("%Y")
    day = now.strftime("%d")
    month = now.strftime("%m")
    month = int(month) - 1

    start_date2 = datetime.strptime(start_date1, "%d/%m/%Y")


    driver.get("https://iiitb.codetantra.com/login.jsp")   # logging in to codetantra
    driver.find_element_by_id("loginId").send_keys(username)
    driver.find_element_by_id("password").send_keys(password)
    driver.find_element_by_id("loginBtn").click()

    driver.implicitly_wait(5)

    # using a try except block to detect if the login was successfull. the try block searches for an element present only on the
    # page after login. If login is not successfull, it raises an exception.
    try:
        driver.find_element(By.XPATH, "//div[@class = 'col my-2 my-md-3 my-lg-4']//div[@class = 'card h-100']//div[@class = 'card-footer p-0']//a[text() = 'Tests']")
        print("Login successful")
    except NoSuchElementException:
        print("Login failed")
        exit()

    classes = driver.find_element(By.XPATH, "//div[@class = 'col my-2 my-md-3 my-lg-4']//div[@class = 'card h-100']//div[@class = 'card-footer p-0']//a[text() = 'View Classes/Meetings']")
    driver.get(classes.get_attribute('href'))

    def list_sorter():
        x = driver.find_elements(By.XPATH, "//th[@class = 'datepicker-switch']")

        x[0].click()
        driver.implicitly_wait(5)
        x[1].click()
        driver.implicitly_wait(5)
        driver.find_element(By.XPATH, "//span[@class = 'year'][text() = '" + start_year + "']").click()
        driver.implicitly_wait(5)
        driver.find_element(By.XPATH, "//span[@class = 'month'][text() = '" + start_date2.strftime("%B")[0:3] + "']").click()
        driver.implicitly_wait(5)
        driver.find_element(By.XPATH, "//td[@class = 'day'][text() = '" + start_day + "']").click()
        driver.implicitly_wait(5)
        WebDriverWait(driver=driver, timeout=10).until(EC.presence_of_element_located((By.XPATH, "//button[@class = 'fc-listView-button fc-button fc-button-primary']")))
        driver.find_element(By.XPATH, "//button[@class = 'fc-listView-button fc-button fc-button-primary']").click()

    list_sorter()
    driver.implicitly_wait(5)
    y = len(driver.find_elements(By.XPATH, "//tr[@class = 'fc-list-item']"))


    for class_name in range(1, y + 1):
        elem = driver.find_element(By.XPATH, "//tr[@class = 'fc-list-item'][" + str(class_name) + "]")
        name = elem.get_attribute('title')
        elem.click()
        try:
            recording = driver.find_element(By.XPATH, "//a[@class = 'text-success']").get_attribute('href')
        except NoSuchElementException:
            recording = "Recording Still Processing"
        topic = input("enter the topic for " + name + " : ")

        names.append(name)
        topics.append(topic)
        recordings.append(recording)
        dates.append(start_date1)
        driver.back()
        list_sorter()


    df = pd.DataFrame()
    df['Class Name'] = names
    df['Date'] = dates
    df['Topic'] = topics
    df['Recording Link'] = recordings

    filename = r"rec.xlsx"

    def append_df_to_excel(filename, df, sheet_name, startrow=None,
                           truncate_sheet=False,
                           **to_excel_kwargs):
        # Excel file doesn't exist - saving and exiting
        if not os.path.isfile(filename):
            df.to_excel(
                filename,
                sheet_name=sheet_name,
                startrow=startrow if startrow is not None else 0,
                **to_excel_kwargs)
            return

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()

    append_df_to_excel(filename, df, sheet_name='Sheet1', index=False, header = False)

    # writer = pd.ExcelWriter(filename,engine = 'xlsxwriter')
    #
    # writer.book = load_workbook(filename)
    # writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    #
    # # Auto-adjust columns' width
    # for column in df:
    #     column_width = max(df[column].astype(str).map(len).max(), len(column))
    #     col_idx = df.columns.get_loc(column)
    #     writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
    #
    # # writer.save()
    # writer.save()

scrape()
driver.close()












